from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

EXCEL_FILE = "data.xlsx"
HTML_FILE = "index.html"

# Function to get the last state from the Excel file
def get_last_state():
    if not os.path.exists(EXCEL_FILE):
        # Create the file if it doesn't exist
        wb = Workbook()
        ws = wb.active
        ws.append(["Time", "State"])  # Add header
        ws.append(["", "On"])  # Default initial state
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    last_row = ws.max_row
    return ws.cell(row=last_row, column=2).value

# Function to log a new state into the Excel file
def log_state(new_state):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    current_time = datetime.now().strftime("%I:%M:%S %p")  # 12-hour time format
    ws.append([current_time, new_state])
    wb.save(EXCEL_FILE)
    print(f"Logged: {current_time}, {new_state}")

# Function to update the HTML file with the current state
def update_html(state):
    with open(HTML_FILE, "r") as file:
        lines = file.readlines()

    updated_lines = []
    for line in lines:
        if "id=\"abdullah\"" in line:
            if state == "On":
                updated_lines.append('        <button id="abdullah" style="background-color: green; color: white;">Abdullah</button>\n')
            else:
                updated_lines.append('        <button id="abdullah" style="background-color: red; color: white;">Abdullah</button>\n')
        else:
            updated_lines.append(line)

    with open(HTML_FILE, "w") as file:
        file.writelines(updated_lines)
    print(f"Updated HTML file with state: {state}")

# Main function to toggle the state and log if changed
def main():
    last_state = get_last_state()

    # Toggle the state
    new_state = "Off" if last_state == "On" else "On"

    if new_state != last_state:
        log_state(new_state)

    # Update the HTML file with the new state
    update_html(new_state)

if __name__ == "__main__":
    main()
